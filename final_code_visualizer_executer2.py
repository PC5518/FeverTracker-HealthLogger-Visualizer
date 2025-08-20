import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
# i used Google AI studio in order to define Comments and take some integrating pandas module logic that i wasn't getting . it helped me a lot
# --- Configuration ---
FILENAME = "fever_tracker.xlsx"
COLUMNS = ["Date & Time", "Temperature (°F)", "Feeling", "Medicine", "Additional Notes"]

def run_data_entry():
    """
    Handles user input for a new health log entry and saves it to the Excel file.
    """
    print("--- New Fever Log Entry ---")
    
    try:
        temperature_f = float(input("🌡️ Enter your current temperature (°F): "))
    except ValueError:
        print("❌ Invalid input. Please enter a number for temperature.")
        return

    feeling = input("🧠 How are you feeling? (e.g., tired, headache) [Press Enter to skip]: ").strip()
    medicine = input("💊 Any medicine taken? [Press Enter to skip]: ").strip()
    note = input("📝 Any additional comment? [Press Enter to skip]: ").strip()

    now = datetime.now()
    timestamp = now.strftime("%d-%b-%Y %I:%M %p")

    new_row = pd.DataFrame(
        [[timestamp, temperature_f, feeling, medicine, note]],
        columns=COLUMNS
    )

    if os.path.exists(FILENAME):
        try:
            old_data = pd.read_excel(FILENAME)
            full_data = pd.concat([old_data, new_row], ignore_index=True)
        except Exception as e:
            print(f"Error reading existing Excel file: {e}. Starting fresh.")
            full_data = new_row
    else:
        full_data = new_row

    full_data.to_excel(FILENAME, index=False)

    wb = load_workbook(FILENAME)
    ws = wb.active

    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for col_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max_length + 4

    wb.save(FILENAME)

    print(f"\n✅ Logged at {timestamp} into '{FILENAME}' successfully.")


def plot_fever_chart():
    """
    Reads data from Excel, plots the trend with zones, and adds a summary
    box for the latest temperature reading.
    """
    print(f"📈 Generating plot from '{FILENAME}'...")

    if not os.path.exists(FILENAME):
        print("File not found. Please add at least one entry first.")
        return

    df = pd.read_excel(FILENAME)

    if df.empty or 'Temperature (°F)' not in df.columns:
        print("No data available to plot.")
        return

    df['Date & Time'] = pd.to_datetime(df['Date & Time'], format='%d-%b-%Y %I:%M %p')
    df['Temperature (°F)'] = pd.to_numeric(df['Temperature (°F)'], errors='coerce')
    df.sort_values('Date & Time', inplace=True)
    
    # --- Define Fever & Plot Boundaries ---
    NORMAL_END = 99.1
    LOW_GRADE_START, LOW_GRADE_END = 99.1, 100.4
    MODERATE_GRADE_START, MODERATE_GRADE_END = 100.6, 102.2
    HIGH_GRADE_START, HIGH_GRADE_END = 102.4, 105.8
    Y_AXIS_MIN, Y_AXIS_MAX = 97.0, 106.0
    
    # --- Plotting ---
    plt.style.use('seaborn-v0_8-whitegrid')
    fig, ax = plt.subplots(figsize=(14, 8))

    ax.set_ylim(Y_AXIS_MIN, Y_AXIS_MAX)
    
    # --- Add Shaded Regions for Fever Levels ---
    ax.axhspan(HIGH_GRADE_START, HIGH_GRADE_END, color="#d83835", alpha=0.7, label=f'High-Grade ({HIGH_GRADE_START}°F - {HIGH_GRADE_END}°F)')
    ax.axhspan(MODERATE_GRADE_START, MODERATE_GRADE_END, color="#e6932d", alpha=0.7, label=f'Moderate-Grade ({MODERATE_GRADE_START}°F - {MODERATE_GRADE_END}°F)')
    ax.axhspan(LOW_GRADE_START, LOW_GRADE_END, color="#f1e727", alpha=0.7, label=f'Low-Grade ({LOW_GRADE_START}°F - {LOW_GRADE_END}°F)')
    ax.axhspan(Y_AXIS_MIN, NORMAL_END, color="#4bda17", alpha=0.7, label=f'Normal (< {NORMAL_END}°F)')

    # --- Plot the Main Temperature Data ---
    ax.plot(
        df['Date & Time'], df['Temperature (°F)'], 
        marker='o', linestyle='-', color='darkblue', 
        linewidth=2.5, markersize=7, label='Your Temperature', zorder=10 
    )

    # --- NEW: Add the Latest Temperature Info Box ---
    if not df.empty:
        # Get the last row of the sorted dataframe
        latest_entry = df.iloc[-1]
        latest_temp = latest_entry['Temperature (°F)']
        latest_timestamp = latest_entry['Date & Time']
        
        # Format the text as requested
        date_str = latest_timestamp.strftime('%d-%b-%Y')
        time_str = latest_timestamp.strftime('%I:%M %p')
        info_text = f"The latest temperature on {date_str} at {time_str} is: {latest_temp}°F"
        
        # Define properties for the curved box
        bbox_props = dict(boxstyle='round,pad=0.5', facecolor='aliceblue', alpha=0.8, edgecolor='black')
        
        # Add the text to the plot in the top-left corner
        ax.text(0.02, 0.98, info_text, 
                transform=ax.transAxes, 
                fontsize=11,
                fontweight='bold',
                verticalalignment='top', 
                bbox=bbox_props)

    # --- Formatting and Labels ---
    ax.set_title('Fever Temperature Graph', fontsize=18, weight='bold', pad=20)
    ax.set_xlabel('Date & Time', fontsize=12)
    ax.set_ylabel('Temperature (°F)', fontsize=12)
    
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%d-%b\n%I:%M %p'))
    fig.autofmt_xdate(rotation=0, ha='center')

    ax.grid(True, which='major', linestyle='--', linewidth=0.5)

    handles, labels = ax.get_legend_handles_labels()
    order = [4, 3, 2, 1, 0] 
    ax.legend([handles[idx] for idx in order], [labels[idx] for idx in order], loc='upper right', title="Legend", fontsize=10)

    plt.tight_layout(rect=[0, 0, 1, 0.96]) # Adjust layout to make space for the title
    plt.show()


if __name__ == "__main__":
    run_data_entry()

    plot_fever_chart()
